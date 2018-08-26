#-*- coding:utf-8 -*-
# 将excel中的电视频道提取关键字并重命名
# 每次只需要修改主程序的path，sheet，getUser方法的userCol和proCol和allUser的保存列数

import re
import sys
import openpyxl
import time

class reName:
    def __init__(self, path, sheet):
        self.path = path
        self.sheet = sheet
        #用户ID和节目名称列号
        #self.userCol = 1
        self.proCol = 4
        self.writeCol = "I"
        # 打开文件
        self.data = openpyxl.load_workbook(path)
        # 打开需要操作的sheet
        self.sheet_name = self.data[sheet]
        self.n_of_rows = self.sheet_name.max_row  # 获取行号

    #获取用户信息
    def getUser(self, row):
        #self.user = self.sheet_name.cell(row=row + 1, column=self.userCol).value
        self.program = self.sheet_name.cell(row=row + 1, column=self.proCol).value

    #提取关键字
    def extractWord(self, row):
        info = re.sub(u"\\（.*?）|\\(.*?\)|\\{.*?}|\\[.*?]|\\【.*?】|\\<.*?>|\\《.*?》", " ", self.program.decode())
        info = info.rstrip()
        # 每一个空格分开成字符串
        words = re.split('[ ]', info)
        new_words = ""
        for word in words:
            #去掉日期信息：有月有日，且字符长度为6即为日期
            if u"月" in word and u"日" in word and len(word) == 6:
                print u"删除日期"
            else:
                new_words += " " + word
        new_words = new_words.lstrip()
        new_info = ""
        D = {}
        for ch in new_words:
            #判断字符不为：或者-时添加字符进入字符串
            if ch == ":" or ch == "：":
                break
            elif ch == "-":
                if ch in D:
                    break
                else:
                    D["-"] = 0
                    new_info = new_info + ch
            else:
                new_info = new_info + ch
        print u"正在提取第" + str(row + 1) + u"个关键字"
        print new_info
        return new_info

    #所有用户数据并保存
    def allUser(self):
        for row in range(self.n_of_rows):
            if row == 0:
                #写入H列
                self.sheet_name[self.writeCol + str(row + 1)] = u"购买套餐"
                print u"开始提取信息"
            else:
                self.getUser(row)
                self.sheet_name[self.writeCol + str(row + 1)] = self.extractWord(row)
        self.data.save(self.path)

reload(sys)
sys.setdefaultencoding('utf-8')
start = time.clock()
#设置路径
path = u'F:\\postgraduate\\期刊论文\\用户信息.xlsx'
sheet = u"Sheet1"
rename = reName(path, sheet)
rename.allUser()
end = time.clock()
print u"完成时长：" + str(end-start)