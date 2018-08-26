#-*- coding:utf-8 -*-
# 将excel中的节目按名称在豆瓣电影上进行爬取
# 每次只需要修改主程序的path，sheet，getUser方法的userCol和proCol和allUser的保存列数

import sys
import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
from requests.exceptions import ReadTimeout,ConnectionError,RequestException
import time
from selenium.common.exceptions import NoSuchElementException

class Spider:
    def __init__(self, path, sheet):
        self.path = path
        self.sheet = sheet
        # 节目名称列号
        self.proCol = 1
        # 打开文件
        self.data = openpyxl.load_workbook(path)
        # 打开需要操作的sheet
        self.sheet_name = self.data[sheet]
        self.n_of_rows = self.sheet_name.max_row  # 获取行号
        #定义需要采集的信息内容
        self.director = u"导演"
        self.screenwriter = u"编剧"
        self.actors = u"主演"
        self.types = u"类型"
        self.area = u"制片"
        self.language = u"语言"
        self.duration = u"片长"
        self.episode = u"集数"
        self.introduction = u"简介"
        self.name = u"名称"
        self.year = u"年份"
        self.rating = u"评分"
        self.num = u"评价人数"
        self.pic = u"图片"


    #获得对应行的节目信息
    def getPro(self, row):
        self.program = self.sheet_name.cell(row=row + 1, column=self.proCol).value

    #用模拟浏览器的方式动态加载网页，将网页的所有电影名称以及网址加入列表
    def getName(self):
        url = "https://movie.douban.com/"
        try:
            # 这个路径就是你添加到PATH的路径
            driver = webdriver.PhantomJS(executable_path='C:/Python27/Scripts/phantomjs-2.1.1-windows/bin/phantomjs.exe')
            driver.get(url)
            # 在搜索框上模拟输入信息并点击
            elem = driver.find_element_by_name("search_text")
            elem.send_keys(self.program)
            elem.send_keys(Keys.RETURN)
            # 得到动态加载的网页
            data = driver.page_source
            soup = BeautifulSoup(data, "lxml")
            l = []
            # 进行匹配
            if soup.select("div[class='item-root']"):
                for i in soup.select("div[class='item-root']"):
                    name = i.find("a", class_="title-text").text
                    name = re.sub(u"\\（.*?）|\\(.*?\)|\\{.*?}|\\[.*?]|\\【.*?】|\\<.*?>|\\《.*?》", " ", name.decode())
                    words = re.split('[ ]', name)
                    if len(words) == 1:
                        rel_name = words[0]
                    else:
                        rel_name = words[0] + words[1]
                    url = i.find("a").get('href')
                    #将得到的名称和网址添加到列表中
                    l.append([rel_name, url])
            driver.quit()
            return l
        except NoSuchElementException as msg:
            print u"查找元素异常%s" % msg

    #根据名称与关键字的相似度，选择需要爬取的网址（最小编辑距离）
    def selectURL(self, name_list = []):
        web = ""
        N = len(self.program.decode())
        # print "N:"+str(N)
        # 设置一个较大的字符相似度
        D_min = 200
        for item in name_list:
            M = len(item[0].decode())
            # print "M:"+str(M)
            D = [[0 for col in range(N + 1)] for row in range(M + 1)]
            i = 0
            for X in item[0].decode():
                i += 1
                D[i][0] = i
                j = 0
                for Y in self.program.decode():
                    j += 1
                    D[0][j] = j
                    D1 = D[i - 1][j] + 1
                    D2 = D[i][j - 1] + 1
                    if X == Y:
                        D3 = D[i - 1][j - 1]
                    else:
                        D3 = D[i - 1][j - 1] + 2
                    D[i][j] = min(D1, D2, D3)
            if D[M][N] < D_min:
                D_min = D[M][N]
                web = item[1]
            # print D[M][N]
        #返回进行处理的页面
        print "web:"+web
        return web

    #根据筛选出的网址进行信息爬取
    def getInfo(self,url):
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36'}
        # proxies = {
        #     "https": "http://119.28.194.66:8888","https": "http://115.198.38.58:6666","https": "http://180.118.240.12:61234",
        #     "https": "http://121.61.89.38:61234","https": "http://112.67.32.246:8118","https": "http://111.76.137.119:808"
        # }
        sleep_download_time = 10
        # 先将键值初始化
        director_value, screenwriter_value, actors_value, types_value, area_value, language_value = "", "", "", "", "", ""
        duration_value, episode_value, introduction_value, rating_value, num_value, year_value = "", "", "", "", "", ""
        pic_value, name_value, info = "", "", ""
        try:
            time.sleep(sleep_download_time)
            response = requests.get(url, headers=headers)
            page = response.text
            soup = BeautifulSoup(page, "lxml")
            if soup.find('div', {'id': 'info'}):
                info = soup.find('div', {'id': 'info'}).text
                words = re.split('[\t\n\r]', info)
                for word in words:
                    if self.director in word:
                        key = re.split('[:]', word)
                        director_value = key[1].encode("gbk", "ignore")
                        continue
                    elif self.screenwriter in word:
                        key = re.split('[:]', word)
                        screenwriter_value = key[1].encode("gbk", "ignore")
                        continue
                    elif self.actors in word:
                        key = re.split('[:]', word)
                        actors_value = key[1].encode("gbk", "ignore")
                        continue
                    elif self.types in word:
                        key = re.split('[:]', word)
                        types_value = key[1].encode("gbk", "ignore")
                        continue
                    elif self.area in word:
                        key = re.split('[:]', word)
                        area_value = key[1].encode("gbk", "ignore")
                        continue
                    elif self.language in word:
                        key = re.split('[:]', word)
                        language_value = key[1].encode("gbk", "ignore")
                        continue
                    elif self.duration in word:
                        # 找到时长参数里面的数字部分，多个时长只取第一个
                        duration_value = re.findall(r"\d+", word)[0]
                        continue
                    elif self.episode in word:
                        key = re.split('[:]', word)
                        episode_value = key[1].encode("gbk", "ignore")
                        continue
            if soup.find('span', {'property': 'v:itemreviewed'}):
                name_value = soup.find('span', {'property': 'v:itemreviewed'}).text.encode('gbk', 'ignore')
            if soup.find('span', {'class': 'year'}):
                year_value = re.findall(r"\d+", soup.find('span', {'class': 'year'}).text)[0]
            if soup.find('span', {'property': 'v:summary'}):
                introduction_value = soup.find('span', {'property': 'v:summary'}).text.encode('gbk', 'ignore')
            if soup.find('strong', {'property': 'v:average'}):
                rating_value = soup.find('strong', {'property': 'v:average'}).text
            if soup.find('span', {'property': 'v:votes'}):
                num_value = soup.find('span', {'property': 'v:votes'}).text.encode('gbk', 'ignore')
            if soup.find('img', {'rel': 'v:image'}):
                pic_value = soup.find('img', {'rel': 'v:image'}).get('src')
        except ReadTimeout:
            print("timeout")
        except ConnectionError:
            print("connection Error")
        except RequestException:
            print("error")

        return [name_value,year_value,introduction_value,rating_value,num_value,director_value,screenwriter_value,
                actors_value,area_value,language_value,duration_value,episode_value,types_value,pic_value]

    # 保存信息到excel
    def saveInfo(self):
        col = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O']
        col_name = [self.name,self.year,self.introduction,self.rating,self.num,self.director,self.screenwriter,
                    self.actors,self.area,self.language,self.duration,self.episode,self.types,self.pic]
        col_num = len(col)
        for row in range(500,self.n_of_rows):
            if row == 0:
                #写入列
                for item in range(col_num):
                    self.sheet_name[col[item] + str(row + 1)] = col_name[item]
                print u"开始提取信息"
            else:
                self.getPro(row)
                name_list = self.getName()
                if name_list == []:
                    print u"查不到此数据"
                else:
                    l = self.getInfo(self.selectURL(name_list))
                    for item in range(col_num):
                        self.sheet_name[col[item] + str(row + 1)] = l[item].lstrip().rstrip().decode('gbk')
                    print u"查找第" + str(row) + u"条记录完成"
        self.data.save(self.path)


# 设置编码
reload(sys)
sys.setdefaultencoding('utf-8')
# 获得系统编码格式
type = sys.getfilesystemencoding()
start =time.clock()
path = u'F:\\postgraduate\\期刊论文\\节目信息.xlsx'
sheet = u"Sheet1"
spider = Spider(path, sheet)
spider.saveInfo()
end = time.clock()
print u"任务已完成"
print u"完成时长：" + str(end-start)